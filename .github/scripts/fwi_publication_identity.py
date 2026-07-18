#!/usr/bin/env python3
"""Build and apply the FWI-PUB-ID-1.0 publication identity package.

The controlled manifest is the implementation source of truth. Building the
manifest requires the approved inventory workbook. Applying or checking the
manifest uses only the Python standard library so it can run in CI.
"""

from __future__ import annotations

import argparse
import html as html_lib
import json
import re
import sys
from pathlib import Path
from typing import Any
from urllib.parse import urljoin


METADATA_START = "<!-- FWI-PUBLICATION-METADATA:START -->"
METADATA_END = "<!-- FWI-PUBLICATION-METADATA:END -->"
IDENTITY_START = "<!-- FWI-PUBLICATION-IDENTITY:START -->"
IDENTITY_END = "<!-- FWI-PUBLICATION-IDENTITY:END -->"
RUNTIME_START = "// FWI-PUBLICATION-RUNTIME:START"
RUNTIME_END = "// FWI-PUBLICATION-RUNTIME:END"
RUNTIME_INJECT_START = "// FWI-PUBLICATION-RUNTIME-INJECT:START"
RUNTIME_INJECT_END = "// FWI-PUBLICATION-RUNTIME-INJECT:END"
UPDATED_ISO = "2026-07-15"
UPDATED_HUMAN = "15 July 2026"
DEFAULT_IMAGE = "https://futureworldintelligence.org/assets/logo.png"

FAMILY_NAMES = {
    "Research and strategic analysis": "Research and Strategic Analysis",
    "Briefs and applied knowledge": "Briefs and Applied Knowledge",
    "Principles and explainers": "Principles and Explainers",
    "Courses and learning resources": "Courses and Learning Resources",
    "Multimedia publication": "Multimedia Publications",
    "Institutional publications": "Institutional Publications",
}

DOMAIN_NAMES = {
    "AI": "AI Intelligence",
    "Climate": "Climate Intelligence",
    "Energy": "Energy Intelligence",
    "Futures": "Strategic Futures",
    "Geopolitics": "Geopolitics Intelligence",
    "Education": "Learning and Education",
    "Institutional": "Institutional Governance",
}

SCHEMA_TYPES = {
    "FWI Course Landing Page": "Course",
    "FWI Course Lecture": "LearningResource",
    "FWI Course Roadmap": "LearningResource",
    "FWI Learning Resource / Prompt Library": "LearningResource",
    "FWI Module Guide": "LearningResource",
    "FWI Multimedia Derivative": "VideoObject",
    "FWI Multimedia Intelligence Brief": "VideoObject",
    "FWI Key Principle / Educational Explainer": "Article",
    "FWI Key Principle / Research Explainer": "Article",
    "FWI Intelligence Brief": "Article",
    "FWI Evidence Review / Conceptual Research Report": "ScholarlyArticle",
    "Institutional Policy / Legal Publication": "DigitalDocument",
}

CONTROLLED_META_NAMES = {
    "description",
    "author",
    "publisher",
    "robots",
    "citation_title",
    "citation_author",
    "citation_publication_date",
    "citation_language",
    "citation_public_url",
    "twitter:card",
    "twitter:title",
    "twitter:description",
    "twitter:image",
    "twitter:image:alt",
}


def parse_attrs(tag: str) -> dict[str, str]:
    attrs: dict[str, str] = {}
    pattern = re.compile(r"([:\w-]+)\s*=\s*(?:\"([^\"]*)\"|'([^']*)'|([^\s>]+))")
    for match in pattern.finditer(tag):
        attrs[match.group(1).lower()] = next(
            value for value in match.groups()[1:] if value is not None
        )
    return attrs


def strip_tags(value: str) -> str:
    value = re.sub(r"<script\b[^>]*>.*?</script>", " ", value, flags=re.I | re.S)
    value = re.sub(r"<style\b[^>]*>.*?</style>", " ", value, flags=re.I | re.S)
    value = re.sub(r"<[^>]+>", " ", value)
    return html_lib.unescape(re.sub(r"\s+", " ", value)).strip()


def clean_series(value: str | None, publication_type: str) -> str:
    value = (value or "").strip()
    low = value.lower()
    if not value:
        return "Standalone publication"
    if "currently labelled" in low:
        return "Climate publication — numbering under review"
    if "proposed" in low:
        return re.sub(r"\s*\(proposed\)\s*", " — numbering under review", value, flags=re.I)
    if "provisional" in low:
        return "Geopolitics Key Principles — number pending"
    if low.startswith("unnumbered"):
        return "Standalone " + value[len("Unnumbered ") :]
    if publication_type == "FWI Multimedia Derivative" and value == "Ten Maps video":
        return "Ten Maps video adaptation"
    return value


def schema_type(publication_type: str) -> str:
    if publication_type in SCHEMA_TYPES:
        return SCHEMA_TYPES[publication_type]
    if "Brief" in publication_type:
        return "Article"
    return "Report"


def infer_page_details(path: Path, public_url: str, fallback_title: str) -> tuple[str, str, str | None]:
    text = path.read_text(encoding="utf-8")
    try:
        from lxml import html as lxml_html

        doc = lxml_html.fromstring(text)
        headings = [" ".join(node.itertext()).strip() for node in doc.xpath("//h1")]
        headings = [re.sub(r"\s+", " ", value) for value in headings if value]
        title = headings[0] if headings else fallback_title.split(" | ")[0].strip()
        images = doc.xpath("//meta[translate(@property,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz')='og:image']/@content")
        if not images:
            images = doc.xpath("//img/@src")
        image = urljoin(public_url, images[0]) if images else DEFAULT_IMAGE
        date_values = doc.xpath("//meta[translate(@name,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz')='citation_publication_date']/@content")
        date_values += re.findall(r'"datePublished"\s*:\s*"([^\"]+)', text)
        date_values += [" ".join(node.itertext()) for node in doc.xpath("//footer")]
    except Exception:
        match = re.search(r"<h1\b[^>]*>(.*?)</h1>", text, flags=re.I | re.S)
        title = strip_tags(match.group(1)) if match else fallback_title.split(" | ")[0].strip()
        image_match = re.search(r"<meta\b[^>]*property=[\"']og:image[\"'][^>]*content=[\"']([^\"']+)", text, flags=re.I)
        image = urljoin(public_url, image_match.group(1)) if image_match else DEFAULT_IMAGE
        date_values = re.findall(r'"datePublished"\s*:\s*"([^\"]+)', text)
        date_values += [strip_tags(value) for value in re.findall(r"<footer\b[^>]*>(.*?)</footer>", text, flags=re.I | re.S)]

    year = None
    for value in date_values:
        match = re.search(r"20(?:1\d|2\d)", str(value))
        if match:
            year = match.group(0)
            break
    return title, image, year


def build_manifest(workbook_path: Path, repo_root: Path, manifest_path: Path) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("Building the manifest requires openpyxl.") from exc

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    register_sheet = workbook["Publication Register"]
    register_headers = [cell.value for cell in next(register_sheet.iter_rows(min_row=4, max_row=4))]
    register: dict[str, dict[str, Any]] = {}
    for values in register_sheet.iter_rows(min_row=5, values_only=True):
        row = dict(zip(register_headers, values))
        if row.get("Inventory ID"):
            register[str(row["Inventory ID"])] = row

    category_sheet = workbook["Category Stratification"]
    category_headers = None
    primary_rows: list[dict[str, Any]] = []
    for values in category_sheet.iter_rows(values_only=True):
        if values and values[0] == "Governance Stratum" and "Source Path" in values:
            category_headers = list(values)
            continue
        if category_headers and values and values[0] == "Primary publication":
            primary_rows.append(dict(zip(category_headers, values)))

    publications: list[dict[str, Any]] = []
    for classification in primary_rows:
        inventory_id = str(classification["Inventory ID"])
        row = register[inventory_id]
        source_path = str(row["Source Path"])
        page_path = repo_root / source_path
        if not page_path.is_file():
            raise SystemExit(f"Inventory path does not exist: {source_path}")
        public_url = str(row["Public URL"])
        title, image, year = infer_page_details(page_path, public_url, str(row["Current Title"]))
        publication_type = str(row["Recommended Classification"])
        publications.append(
            {
                "inventory_id": inventory_id,
                "source_path": source_path,
                "public_url": public_url,
                "family": FAMILY_NAMES[str(classification["Publication Family"])],
                "publication_type": publication_type,
                "domain": DOMAIN_NAMES.get(str(row["Domain"]), str(row["Domain"])),
                "series": clean_series(row.get("Current Series / Label"), publication_type),
                "title": title,
                "purpose": str(row["Purpose"]),
                "audience": str(row["Primary Audience"]),
                "method": str(row["Method / Evidence Basis"]),
                "publication_year": year,
                "schema_type": schema_type(publication_type),
                "social_image": image,
            }
        )

    publications.sort(key=lambda item: item["source_path"].lower())
    if len(publications) != 72:
        raise SystemExit(f"Expected 72 primary publications; found {len(publications)}")
    payload = {
        "standard": "FWI-PUB-ID-1.0",
        "status": "Approved standard — controlled implementation in progress",
        "metadata_updated": UPDATED_ISO,
        "source_inventory": workbook_path.name,
        "primary_publication_count": 72,
        "publications": publications,
    }
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def remove_controlled_metadata(text: str) -> str:
    text = re.sub(
        re.escape(METADATA_START) + r".*?" + re.escape(METADATA_END) + r"\s*",
        "",
        text,
        flags=re.S,
    )

    def meta_replacement(match: re.Match[str]) -> str:
        attrs = parse_attrs(match.group(0))
        name = attrs.get("name", "").lower()
        prop = attrs.get("property", "").lower()
        if name in CONTROLLED_META_NAMES or prop.startswith("og:") or prop == "article:modified_time":
            return ""
        return match.group(0)

    text = re.sub(r"<meta\b[^>]*>", meta_replacement, text, flags=re.I)

    def link_replacement(match: re.Match[str]) -> str:
        attrs = parse_attrs(match.group(0))
        rel = attrs.get("rel", "").lower().split()
        if (
            "canonical" in rel
            or attrs.get("data-fwi-publication-style") == "1"
            or attrs.get("data-fwi-hero-style") == "1"
        ):
            return ""
        return match.group(0)

    return re.sub(r"<link\b[^>]*>", link_replacement, text, flags=re.I)


def metadata_block(item: dict[str, Any]) -> str:
    esc = html_lib.escape
    title = esc(item["title"], quote=True)
    description = esc(item["purpose"], quote=True)
    url = esc(item["public_url"], quote=True)
    image = esc(item["social_image"], quote=True)
    year = item.get("publication_year")
    og_type = "video.other" if item["schema_type"] == "VideoObject" else "article"

    json_ld: dict[str, Any] = {
        "@context": "https://schema.org",
        "@type": item["schema_type"],
        "name": item["title"],
        "headline": item["title"],
        "description": item["purpose"],
        "url": item["public_url"],
        "image": item["social_image"],
        "inLanguage": "en",
        "dateModified": UPDATED_ISO,
        "version": "Web edition 1.0",
        "creativeWorkStatus": "Published; retrospective validation pending",
        "author": {"@type": "Organization", "name": "FutureWorld Intelligence"},
        "publisher": {"@type": "Organization", "name": "FutureWorld Intelligence"},
        "isPartOf": {"@type": "CreativeWorkSeries", "name": item["series"]},
        "audience": {"@type": "Audience", "audienceType": item["audience"]},
        "about": item["domain"],
    }
    if year:
        json_ld["datePublished"] = year

    lines = [
        METADATA_START,
        '<link rel="stylesheet" href="/style.css?v=3.2" data-fwi-publication-style="1" />',
        '<link rel="stylesheet" href="/assets/fwi-universal-hero.css?v=1.0" data-fwi-hero-style="1" />',
        f'<meta name="description" content="{description}" />',
        '<meta name="author" content="FutureWorld Intelligence" />',
        '<meta name="publisher" content="FutureWorld Intelligence" />',
        '<meta name="robots" content="index,follow,max-image-preview:large" />',
        f'<meta name="citation_title" content="{title}" />',
        '<meta name="citation_author" content="FutureWorld Intelligence" />',
    ]
    if year:
        lines.append(f'<meta name="citation_publication_date" content="{esc(str(year), quote=True)}" />')
    lines += [
        '<meta name="citation_language" content="en" />',
        f'<meta name="citation_public_url" content="{url}" />',
        f'<link rel="canonical" href="{url}" />',
        f'<meta property="og:type" content="{og_type}" />',
        '<meta property="og:site_name" content="FutureWorld Intelligence" />',
        f'<meta property="og:title" content="{title}" />',
        f'<meta property="og:description" content="{description}" />',
        f'<meta property="og:url" content="{url}" />',
        f'<meta property="og:image" content="{image}" />',
        f'<meta property="og:image:alt" content="{title} — FutureWorld Intelligence" />',
        '<meta property="og:locale" content="en_US" />',
        f'<meta property="article:modified_time" content="{UPDATED_ISO}" />',
        '<meta name="twitter:card" content="summary_large_image" />',
        f'<meta name="twitter:title" content="{title}" />',
        f'<meta name="twitter:description" content="{description}" />',
        f'<meta name="twitter:image" content="{image}" />',
        f'<meta name="twitter:image:alt" content="{title} — FutureWorld Intelligence" />',
        '<script type="application/ld+json">',
        json.dumps(json_ld, ensure_ascii=False, indent=2).replace("</", r"<\/"),
        "</script>",
        METADATA_END,
    ]
    return "\n".join(lines)


def evidence_cutoff(item: dict[str, Any]) -> str:
    family = item["family"]
    if family == "Courses and Learning Resources":
        return "Not applicable as a research evidence cut-off. This learning resource is subject to periodic instructional and technical review."
    if family == "Institutional Publications":
        return "Not applicable as a research evidence cut-off. This policy page reflects the current web edition and requires maintained effective-date and change-history records."
    return "The exact historical evidence cut-off was not recorded when the original web publication was prepared. Source currency will be confirmed during retrospective validation."


def limitation_text(item: dict[str, Any]) -> str:
    family = item["family"]
    if family == "Courses and Learning Resources":
        return "Educational guidance only. Tools, interfaces and procedures may change; instructional accuracy, accessibility, rights and AI-use review remain part of the pending retrospective validation."
    if family == "Multimedia Publications":
        return "This multimedia publication may summarize or adapt source material. Source consistency, media rights, accessibility and disclosure checks remain part of the pending retrospective validation."
    if family == "Institutional Publications":
        return "This page states FutureWorld Intelligence institutional policy for the current web edition. Effective dates, applicability and change history should be confirmed before formal reliance."
    return "Classification and metadata do not independently validate substantive claims. Citation, factual, originality, AI-use, rights and conflict-of-interest checks remain part of the pending retrospective validation."


def identity_block(item: dict[str, Any]) -> str:
    esc = html_lib.escape
    year = str(item["publication_year"]) if item.get("publication_year") else "Not recorded"
    citation_year = str(item["publication_year"]) if item.get("publication_year") else "n.d."
    series = item["series"]
    citation_series = "" if series == "Standalone publication" else f" ({esc(series)}; Web edition 1.0)"
    multimedia_class = " publication-identity--multimedia" if item["family"] == "Multimedia Publications" or item["publication_type"] == "FWI Field/GIS Assessment Presentation" else ""
    fields = [
        ("Publication family", item["family"], False),
        ("Publication type", item["publication_type"], False),
        ("Domain", item["domain"], False),
        ("Series and number", series, False),
        ("Institutional author", "FutureWorld Intelligence", False),
        ("Publication year", year, False),
        ("Current web edition", "1.0", False),
        ("Metadata updated", UPDATED_HUMAN, False),
        ("Purpose", item["purpose"], True),
        ("Intended audience", item["audience"], True),
        ("Method and evidence basis", item["method"], True),
        ("Evidence cut-off", evidence_cutoff(item), True),
        ("Limitations and disclosures", limitation_text(item), True),
    ]
    field_html = "".join(
        f'<div class="publication-identity__item{" publication-identity__item--wide" if wide else ""}"><dt>{esc(label)}</dt><dd>{esc(str(value))}</dd></div>'
        for label, value, wide in fields
    )
    url = esc(item["public_url"], quote=True)
    title = esc(item["title"])
    return (
        f"{IDENTITY_START}\n"
        f'<section class="publication-identity{multimedia_class}" id="publication-details" aria-labelledby="publication-identity-heading">'
        '<div class="publication-identity__inner"><div class="publication-identity__header"><div>'
        '<p class="publication-identity__eyebrow">FWI publication information</p>'
        '<h2 id="publication-identity-heading">Identity, scope and status</h2></div>'
        '<span class="publication-status">Retrospective validation pending</span></div>'
        f'<dl class="publication-identity__grid">{field_html}</dl>'
        '<p class="publication-identity__note"><strong>Validation note:</strong> This classification does not itself validate the publication. Retrospective factual, citation, originality, disclosure and readiness checks must be completed and human-approved before the status can change to “Validated — human approved.”</p>'
        '<div class="publication-citation" aria-label="Recommended citation"><strong>Recommended citation</strong>'
        f'<p>FutureWorld Intelligence. ({citation_year}). <em>{title}</em>{citation_series}. '
        f'<a href="{url}">{url}</a></p></div></div></section>\n'
        f"{IDENTITY_END}"
    )


def add_body_classes(text: str, multimedia: bool) -> str:
    def replacement(match: re.Match[str]) -> str:
        tag = match.group(0)
        required = ["fwi-has-publication-identity"]
        if multimedia:
            required.append("fwi-publication-multimedia")
        class_match = re.search(r"\bclass\s*=\s*([\"'])(.*?)\1", tag, flags=re.I | re.S)
        if class_match:
            classes = class_match.group(2).split()
            for value in required:
                if value not in classes:
                    classes.append(value)
            new_attr = f'class="{" ".join(classes)}"'
            return tag[: class_match.start()] + new_attr + tag[class_match.end() :]
        return tag[:-1] + f' class="{" ".join(required)}">'

    return re.sub(r"<body\b[^>]*>", replacement, text, count=1, flags=re.I)


def add_hero_heading_classes(text: str, multimedia: bool) -> str:
    """Attach the universal hero contract directly to every publication H1."""

    def replacement(match: re.Match[str]) -> str:
        tag = match.group(0)
        required = ["fwi-universal-hero-title"]
        if multimedia:
            required.append("fwi-universal-hero-title--multimedia")
        class_match = re.search(r"\bclass\s*=\s*([\"'])(.*?)\1", tag, flags=re.I | re.S)
        if class_match:
            classes = class_match.group(2).split()
            for value in required:
                if value not in classes:
                    classes.append(value)
            new_attr = f'class="{" ".join(classes)}"'
            return tag[: class_match.start()] + new_attr + tag[class_match.end() :]
        return tag[:-1] + f' class="{" ".join(required)}">'

    return re.sub(r"<h1\b[^>]*>", replacement, text, flags=re.I)


def insert_identity(text: str, block: str, multimedia: bool) -> str:
    if multimedia:
        body_end = re.search(r"</body\s*>", text, flags=re.I)
        if not body_end:
            raise ValueError("missing </body>")
        return text[: body_end.start()].rstrip() + "\n" + block + "\n" + text[body_end.start() :]

    main_match = re.search(r"<main\b[^>]*>", text, flags=re.I)
    if not main_match:
        body_match = re.search(r"<body\b[^>]*>", text, flags=re.I)
        if not body_match:
            raise ValueError("missing <body> and <main>")
        return text[: body_match.end()] + "\n" + block + "\n" + text[body_match.end() :]

    before_main = text[max(0, main_match.start() - 8000) : main_match.start()]
    header_matches = list(re.finditer(r"<header\b[^>]*class\s*=\s*[\"'][^\"']*hero[^\"']*[\"'][^>]*>", before_main, flags=re.I))
    if header_matches and before_main.lower().rfind("</header>") > header_matches[-1].start():
        return text[: main_match.start()] + block + "\n" + text[main_match.start() :]

    after = text[main_match.end() :]
    section_match = re.search(r"<section\b([^>]*)>", after, flags=re.I)
    if section_match:
        attrs = section_match.group(1).lower()
        if any(token in attrs for token in ("hero", "masthead", "cover")):
            close_match = re.search(r"</section\s*>", after[section_match.end() :], flags=re.I)
            if close_match:
                insertion = main_match.end() + section_match.end() + close_match.end()
                return text[:insertion] + "\n" + block + "\n" + text[insertion:]

    return text[: main_match.end()] + "\n" + block + "\n" + text[main_match.end() :]


def apply_runtime_loader_item(text: str, item: dict[str, Any]) -> str:
    """Inject the identity into the HTML fetched by the legacy #011 loader."""
    text = re.sub(
        re.escape(IDENTITY_START) + r".*?" + re.escape(IDENTITY_END) + r"\s*",
        "",
        text,
        flags=re.S,
    )
    text = remove_controlled_metadata(text)
    text = re.sub(re.escape(RUNTIME_START) + r".*?" + re.escape(RUNTIME_END) + r"\s*", "", text, flags=re.S)
    text = re.sub(
        re.escape(RUNTIME_INJECT_START) + r".*?" + re.escape(RUNTIME_INJECT_END) + r"\s*",
        "",
        text,
        flags=re.S,
    )

    # Repair the fallback string if a prior generic insertion placed raw HTML
    # inside its single-quoted JavaScript string.
    fallback = (
        ".catch(()=>{document.body.innerHTML='<main style=\"background:#030712;color:white;"
        "font-family:Arial;padding:40px\">'+fwiPublicationIdentity+"
        "'<h1>Geopolitics Intelligence #011</h1><p>Report loading failed. Please refresh the page.</p></main>';"
        "document.body.classList.add('fwi-has-publication-identity')})})();"
    )
    text = re.sub(r"\.catch\(\(\)=>\{document\.body\.innerHTML=.*?\}\)\(\);", fallback, text, count=1, flags=re.S)

    runtime_metadata = metadata_block(item).replace("</script>", r"<\/script>")
    runtime_identity = identity_block(item)
    declarations = (
        f"{RUNTIME_START}\n"
        f"const fwiPublicationMetadata={json.dumps(runtime_metadata, ensure_ascii=False)};\n"
        f"const fwiPublicationIdentity={json.dumps(runtime_identity, ensure_ascii=False)};\n"
        f"{RUNTIME_END}\n"
    )
    base_match = re.search(r"const base=.*?;", text)
    if not base_match:
        raise ValueError(f"{item['source_path']}: loader base declaration not found")
    text = text[: base_match.end()].rstrip() + "\n" + declarations + text[base_match.end() :].lstrip("\r\n")

    injection = (
        f"  {RUNTIME_INJECT_START}\n"
        "  h=h.replace(/<\\/head>/i,fwiPublicationMetadata+'\\n</head>');\n"
        "  h=h.replace(/<main\\b[^>]*>/i,m=>m+'\\n'+fwiPublicationIdentity);\n"
        f"  {RUNTIME_INJECT_END}\n"
    )
    if "  return h;" not in text:
        raise ValueError(f"{item['source_path']}: loader return statement not found")
    text = text.replace("  return h;", injection + "  return h;", 1)

    head_matches = list(re.finditer(r"</head\s*>", text, flags=re.I))
    if not head_matches:
        raise ValueError(f"{item['source_path']}: missing </head>")
    head_end = head_matches[-1]
    host_metadata = metadata_block(item)
    text = text[: head_end.start()].rstrip() + "\n" + host_metadata + "\n" + text[head_end.start() :]
    return text


def apply_item(repo_root: Path, item: dict[str, Any]) -> bool:
    path = repo_root / item["source_path"]
    original = path.read_text(encoding="utf-8")
    text = original
    has_approved_pilot = (
        item["inventory_id"] == "FWI-INV-028"
        and 'class="publication-identity"' in text
        and IDENTITY_START not in text
    )
    multimedia = item["family"] == "Multimedia Publications" or item["publication_type"] == "FWI Field/GIS Assessment Presentation"

    is_runtime_loader = "document.write(clean(h))" in text and "raw.githubusercontent.com" in text
    if is_runtime_loader:
        text = apply_runtime_loader_item(text, item)
        text = add_body_classes(text, multimedia)
        text = add_hero_heading_classes(text, multimedia)
        if text != original:
            path.write_text(text, encoding="utf-8")
            return True
        return False

    if not has_approved_pilot:
        text = re.sub(
            re.escape(IDENTITY_START) + r".*?" + re.escape(IDENTITY_END) + r"\s*",
            "",
            text,
            flags=re.S,
        )
        text = remove_controlled_metadata(text)
        head_matches = list(re.finditer(r"</head\s*>", text, flags=re.I))
        if not head_matches:
            raise ValueError(f"{item['source_path']}: missing </head>")
        head_end = head_matches[-1]
        meta = metadata_block(item)
        text = text[: head_end.start()].rstrip() + "\n" + meta + "\n" + text[head_end.start() :]
        text = insert_identity(text, identity_block(item), multimedia)

    text = add_body_classes(text, multimedia)
    text = add_hero_heading_classes(text, multimedia)
    if text != original:
        path.write_text(text, encoding="utf-8")
        return True
    return False


def validate(repo_root: Path, manifest: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    publications = manifest.get("publications", [])
    if len(publications) != 72:
        errors.append(f"manifest contains {len(publications)} publications, expected 72")
    paths = [item["source_path"] for item in publications]
    if len(paths) != len(set(paths)):
        errors.append("manifest contains duplicate source paths")
    urls = [item["public_url"] for item in publications]
    if len(urls) != len(set(urls)):
        errors.append("manifest contains duplicate canonical URLs")

    for item in publications:
        multimedia = item["family"] == "Multimedia Publications" or item["publication_type"] == "FWI Field/GIS Assessment Presentation"
        path = repo_root / item["source_path"]
        if not path.is_file():
            errors.append(f"missing publication: {item['source_path']}")
            continue
        text = path.read_text(encoding="utf-8")
        required = [
            "publication-details",
            "Retrospective validation pending",
            item["family"],
            item["publication_type"],
            item["public_url"],
            "fwi-has-publication-identity",
            'data-fwi-hero-style="1"',
            "fwi-universal-hero-title",
        ]
        for value in required:
            if value not in text and html_lib.escape(str(value)) not in text:
                errors.append(f"{item['source_path']}: missing {value!r}")
        if text.lower().count('rel="canonical"') != 1:
            errors.append(f"{item['source_path']}: expected exactly one canonical link")
        if text.count('data-fwi-hero-style="1"') != 1:
            errors.append(f"{item['source_path']}: expected exactly one universal hero stylesheet link")
        if multimedia and "fwi-universal-hero-title--multimedia" not in text:
            errors.append(f"{item['source_path']}: missing multimedia hero-title variant")
        if "const fwiPublicationIdentity=" in text:
            if text.count("publication-details") != 1:
                errors.append(f"{item['source_path']}: expected one runtime publication-details block")
        elif text.count('id="publication-details"') != 1:
            errors.append(f"{item['source_path']}: expected exactly one publication-details block")
        if item["source_path"].endswith(".html") and not re.search(r"</html\s*>", text, flags=re.I):
            errors.append(f"{item['source_path']}: missing </html>")
    return errors


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--repo", type=Path, default=Path.cwd())
    parser.add_argument("--manifest", type=Path, default=Path("governance/fwi-publication-manifest.json"))
    parser.add_argument("--build-manifest", type=Path, metavar="WORKBOOK")
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--check", action="store_true")
    args = parser.parse_args()

    repo_root = args.repo.resolve()
    manifest_path = args.manifest if args.manifest.is_absolute() else repo_root / args.manifest
    if args.build_manifest:
        build_manifest(args.build_manifest.resolve(), repo_root, manifest_path)
        print(f"Built {manifest_path}")

    if not manifest_path.is_file():
        parser.error(f"manifest not found: {manifest_path}")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))

    if args.apply:
        changed = 0
        for item in manifest["publications"]:
            if apply_item(repo_root, item):
                changed += 1
        print(f"Updated {changed} publication pages")

    if args.check or args.apply:
        errors = validate(repo_root, manifest)
        if errors:
            print("Publication identity validation failed:", file=sys.stderr)
            for error in errors:
                print(f"- {error}", file=sys.stderr)
            return 1
        print("Validated 72/72 primary publication identity packages")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
